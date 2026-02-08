import os
import sys
import logging
from openpyxl import Workbook
from utils import resource_path


def export_for_powerbi(assignments, iso_actions=None, filename="lpa_powerbi.xlsx"):
    try:
        logging.info(f"Starting export with {len(assignments)} assignments")

        if iso_actions is None:
            iso_actions = []

        # Determine base path based on how app is running
        if getattr(sys, 'frozen', False):
            # Running as executable
            base_dir = os.path.dirname(sys.executable)
            output_dir = os.path.join(base_dir, "outputs")
        else:
            # Running as script
            base_dir = os.getcwd()
            output_dir = os.path.join(base_dir, "outputs")

        logging.info(f"Base directory: {base_dir}")
        logging.info(f"Output directory: {output_dir}")

        os.makedirs(output_dir, exist_ok=True)
        logging.info(f"Output directory exists: {os.path.exists(output_dir)}")

        filepath = os.path.join(output_dir, filename)
        logging.info(f"Full file path: {filepath}")

        wb = Workbook()

        # ---------------- LPA SCHEDULE ----------------
        ws = wb.active
        ws.title = "LPA_Schedule"
        ws.append(["Section", "Target_Shift", "Auditors"])

        for a in assignments:
            ws.append([a.section, a.target_shift, ", ".join(a.auditors)])

        # ---------------- LPA SUMMARY ----------------
        ws2 = wb.create_sheet("LPA_Summary")
        ws2.append(["Auditor", "LPA_Count"])

        counts = {}
        for a in assignments:
            for auditor in a.auditors:
                counts[auditor] = counts.get(auditor, 0) + 1

        for auditor, count in counts.items():
            ws2.append([auditor, count])

        # ---------------- ISO ACTIONS ----------------
        ws3 = wb.create_sheet("ISO_Actions")
        ws3.append(["Description", "Owner", "Due_Date", "Status", "Overdue"])

        for a in iso_actions:
            ws3.append([
                a.get("description", ""),
                a.get("owner", ""),
                a.get("due_date", ""),
                a.get("status", ""),
                a.get("overdue", False)
            ])

        wb.save(filepath)

        # Verify file was created
        if os.path.exists(filepath):
            logging.info(f"✅ Excel file successfully created at: {filepath}")
            return filepath
        else:
            logging.error(f"❌ Failed to create file at: {filepath}")
            return None

    except Exception as e:
        logging.error(f"❌ Error in export_for_powerbi: {str(e)}")
        raise